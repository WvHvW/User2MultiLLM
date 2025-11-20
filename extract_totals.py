#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
提取OriginResults.xlsx中所有算法的total_cost值，按分布和算法组织后保存到total.xlsx
"""

import pandas as pd
from openpyxl import load_workbook


def extract_totals():
    """从OriginResults.xlsx提取所有Total行，生成汇总表"""

    # 读取所有sheet名称
    wb = load_workbook('results/OriginResults.xlsx')
    sheet_names = wb.sheetnames

    # 存储结果的字典：{分布名: {算法名: total_cost}}
    results = {}

    # 遍历每个sheet
    for sheet_name in sheet_names:
        df = pd.read_excel('results/OriginResults.xlsx', sheet_name=sheet_name)

        # 筛选出包含"-Total"的行
        total_rows = df[df['algorithm'].str.contains('Total', na=False)]

        # 提取算法名和total_cost
        algorithm_totals = {}
        for _, row in total_rows.iterrows():
            # 移除"-Total"后缀作为算法名
            algorithm_name = row['algorithm'].replace('-Total', '')
            total_cost = row['total_cost']
            algorithm_totals[algorithm_name] = total_cost

        results[sheet_name] = algorithm_totals

    # 转换为DataFrame：行=分布名，列=算法名
    df_result = pd.DataFrame(results).T

    # 确保列按照一定顺序排列（no_split, k-split系列, k-split-augment系列）
    # 提取所有列名并排序
    all_columns = list(df_result.columns)

    # 分组排序
    no_split_cols = [col for col in all_columns if col == 'no_split']
    split_cols = sorted([col for col in all_columns if 'split' in col and 'augment' not in col and col != 'no_split'],
                       key=lambda x: int(x.split('-')[0]) if x.split('-')[0].isdigit() else 0)
    augment_cols = sorted([col for col in all_columns if 'augment' in col],
                         key=lambda x: int(x.split('-')[0]) if x.split('-')[0].isdigit() else 0)

    ordered_columns = no_split_cols + split_cols + augment_cols
    df_result = df_result[ordered_columns]

    # 保存到total.xlsx
    df_result.to_excel('results/total.xlsx', index_label='分布名')

    print(f"已成功提取 {len(sheet_names)} 个分布的总计值")
    print(f"算法数量: {len(df_result.columns)}")
    print(f"结果已保存到: results/total.xlsx")

    return df_result


if __name__ == '__main__':
    df = extract_totals()
    print("\n前5行预览:")
    print(df.head())
